import argparse
import json
import re
import shutil
import sqlite3
import subprocess
import sys
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except Exception:
    openpyxl = None


DEFAULT_PROJECT_DIR = Path(__file__).resolve().parent
DEFAULT_RUNTIME_ROOT = Path(r"C:\AUTOMACAO")
DEFAULT_WORKFLOW_ID = "zN3heKJVLO8w4dG6"


def normalize_number(value: str) -> str:
    digits = re.sub(r"\D", "", str(value or ""))
    if not digits:
        raise ValueError("number_empty")
    return digits


def build_number_variants(number: str) -> list[str]:
    digits = normalize_number(number)
    variants = {digits, f"+{digits}", f"{digits}@s.whatsapp.net", f"{digits}@c.us", f"{digits}@lid"}

    national = digits[2:] if digits.startswith("55") and len(digits) > 10 else digits
    variants.update({national, f"+{national}", f"{national}@s.whatsapp.net", f"{national}@c.us", f"{national}@lid"})

    if len(national) in (10, 11):
        area = national[:2]
        subscriber = national[2:]
        variants.add(f"55{area}{subscriber}")
        variants.add(f"{area}{subscriber}")
        variants.add(f"+55 ({area}) {subscriber}")
        if len(subscriber) == 9 and subscriber.startswith("9"):
            without_mobile_nine = area + subscriber[1:]
            variants.add(without_mobile_nine)
            variants.add(f"55{without_mobile_nine}")
            variants.add(f"+55{without_mobile_nine}")
            variants.add(f"{without_mobile_nine}@s.whatsapp.net")
            variants.add(f"55{without_mobile_nine}@s.whatsapp.net")
            variants.add(f"{without_mobile_nine}@c.us")
            variants.add(f"55{without_mobile_nine}@c.us")
            variants.add(f"{without_mobile_nine}@lid")
            variants.add(f"55{without_mobile_nine}@lid")
        if len(subscriber) == 8:
            variants.add(area + "9" + subscriber)

    return sorted({v for v in variants if v}, key=len, reverse=True)


def build_like_patterns(variants: list[str]) -> list[str]:
    return [f"%{variant}%" for variant in variants if variant]


def contains_any_variant(value: str, variants: list[str]) -> bool:
    text = str(value or "")
    return any(variant in text for variant in variants)


def run(cmd, check=True, capture=False):
    return subprocess.run(
        cmd,
        check=check,
        text=True,
        capture_output=capture,
    )


def ensure_dir(path: Path):
    path.mkdir(parents=True, exist_ok=True)


def backup_file(src: Path, dest_dir: Path, label: str, ts: str):
    if src.exists():
        target = dest_dir / f"{label}-{ts}{src.suffix}"
        shutil.copy2(src, target)
        return str(target)
    return None


def backup_n8n_db(project_dir: Path, backup_dir: Path, ts: str):
    target_name = f"n8n-database-{ts}.sqlite"
    cmd = [
        "docker", "run", "--rm",
        "-v", "ai_n8n_data:/data",
        "-v", f"{backup_dir}:/backup",
        "python:3.11-slim",
        "sh", "-lc",
        f"cp /data/database.sqlite /backup/{target_name}",
    ]
    run(cmd)
    return str(backup_dir / target_name)


def backup_evolution_db(backup_dir: Path, ts: str):
    container_tmp = f"/tmp/evolution-full-{ts}.sql"
    host_target = backup_dir / f"evolution-full-{ts}.sql"
    run([
        "docker", "exec", "evolution-postgres",
        "sh", "-lc",
        f"pg_dump -U evolution -d evolution > {container_tmp}",
    ])
    run(["docker", "cp", f"evolution-postgres:{container_tmp}", str(host_target)])
    return str(host_target)


def get_existing_columns(cur, table: str) -> set[str]:
    return {row[1] for row in cur.execute(f'PRAGMA table_info("{table}")').fetchall()}


def cleanup_sqlite_rows_by_columns(
    db_path: Path,
    table_map: dict[str, list[str]],
    variants: list[str],
    dry_run: bool,
):
    result = {"db": str(db_path), "tables": {}}
    patterns = build_like_patterns(variants)
    conn = sqlite3.connect(db_path)
    cur = conn.cursor()
    for table, candidate_columns in table_map.items():
        existing = [column for column in candidate_columns if column in get_existing_columns(cur, table)]
        if not existing:
            result["tables"][table] = {"matched": 0, "columns": []}
            continue

        where_parts = []
        params = []
        for column in existing:
            for pattern in patterns:
                where_parts.append(f'CAST("{column}" AS TEXT) LIKE ?')
                params.append(pattern)
        where_sql = " OR ".join(where_parts) if where_parts else "1=0"
        count = cur.execute(f'SELECT COUNT(*) FROM "{table}" WHERE {where_sql}', params).fetchone()[0]
        result["tables"][table] = {"matched": count, "columns": existing}
        if not dry_run and count:
            cur.execute(f'DELETE FROM "{table}" WHERE {where_sql}', params)

    if not dry_run:
        conn.commit()

    result["remaining"] = {}
    for table, candidate_columns in table_map.items():
        existing = [column for column in candidate_columns if column in get_existing_columns(cur, table)]
        if not existing:
            result["remaining"][table] = 0
            continue
        where_parts = []
        params = []
        for column in existing:
            for pattern in patterns:
                where_parts.append(f'CAST("{column}" AS TEXT) LIKE ?')
                params.append(pattern)
        where_sql = " OR ".join(where_parts) if where_parts else "1=0"
        result["remaining"][table] = cur.execute(f'SELECT COUNT(*) FROM "{table}" WHERE {where_sql}', params).fetchone()[0]

    conn.close()
    return result


def prune_value(obj, variants: list[str]):
    if isinstance(obj, dict):
        out = {}
        for key, value in obj.items():
            if contains_any_variant(key, variants):
                continue
            pruned = prune_value(value, variants)
            if pruned is None:
                continue
            out[key] = pruned
        return out
    if isinstance(obj, list):
        out = []
        for item in obj:
            pruned = prune_value(item, variants)
            if pruned is None:
                continue
            out.append(pruned)
        return out
    if isinstance(obj, str) and contains_any_variant(obj, variants):
        return None
    return obj


def ensure_staticdata_exclusive_allowlist(static_data: dict, authorized_number: str):
    targets = []
    if isinstance(static_data, dict):
        targets.append(static_data)
        if isinstance(static_data.get("global"), dict):
            targets.append(static_data["global"])
    for target in targets:
        target["alwaysAllowedContacts"] = {
            "numbers": [authorized_number],
            "reason": "controlled_real_test_authorized_only",
            "updatedAt": datetime.utcnow().isoformat() + "Z",
        }


def cleanup_n8n_internal(
    number: str,
    workflow_id: str,
    db_path: str = "/data/database.sqlite",
    dry_run: bool = False,
    authorized_number: str | None = None,
):
    variants = build_number_variants(number)
    patterns = build_like_patterns(variants)
    conn = sqlite3.connect(db_path, timeout=60)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA busy_timeout=60000")
    cur = conn.cursor()

    row = cur.execute("SELECT staticData FROM workflow_entity WHERE id = ?", (workflow_id,)).fetchone()
    static_data_raw = row["staticData"] if row else None
    static_data = json.loads(static_data_raw) if static_data_raw else {}
    static_before_raw = json.dumps(static_data, ensure_ascii=False)
    static_has_number_before = any(variant in static_before_raw for variant in variants)
    pruned_static = prune_value(static_data, variants)
    if not isinstance(pruned_static, dict):
        pruned_static = {}
    if authorized_number:
        ensure_staticdata_exclusive_allowlist(pruned_static, authorized_number)

    execution_where = " OR ".join(["data LIKE ?" for _ in patterns] + ["workflowData LIKE ?" for _ in patterns])
    execution_params = patterns + patterns

    execution_ids = [
        r["executionId"]
        for r in cur.execute(
            f"SELECT executionId FROM execution_data WHERE {execution_where} ORDER BY executionId",
            execution_params,
        ).fetchall()
    ]

    if not dry_run:
        cur.execute(
            'UPDATE workflow_entity SET staticData = ?, updatedAt = STRFTIME("%Y-%m-%d %H:%M:%f", "NOW") WHERE id = ?',
            (json.dumps(pruned_static, ensure_ascii=False, separators=(",", ":")), workflow_id),
        )
        if execution_ids:
            placeholders = ",".join("?" * len(execution_ids))
            annotation_ids = [
                r["id"]
                for r in cur.execute(
                    f"SELECT id FROM execution_annotations WHERE executionId IN ({placeholders})",
                    execution_ids,
                ).fetchall()
            ]
            if annotation_ids:
                ann_placeholders = ",".join("?" * len(annotation_ids))
                cur.execute(
                    f"DELETE FROM execution_annotation_tags WHERE annotationId IN ({ann_placeholders})",
                    annotation_ids,
                )
            cur.execute(f"DELETE FROM execution_annotations WHERE executionId IN ({placeholders})", execution_ids)
            cur.execute(f"DELETE FROM execution_metadata WHERE executionId IN ({placeholders})", execution_ids)
            cur.execute(
                f"DELETE FROM test_case_execution WHERE executionId IN ({placeholders}) OR pastExecutionId IN ({placeholders}) OR evaluationExecutionId IN ({placeholders})",
                execution_ids * 3,
            )
            cur.execute(f"DELETE FROM chat_hub_messages WHERE executionId IN ({placeholders})", execution_ids)
            cur.execute(f"DELETE FROM execution_data WHERE executionId IN ({placeholders})", execution_ids)
            cur.execute(f"DELETE FROM execution_entity WHERE id IN ({placeholders})", execution_ids)
        conn.commit()

    execution_count_after = cur.execute(
        f"SELECT COUNT(*) FROM execution_data WHERE {execution_where}",
        execution_params,
    ).fetchone()[0]
    static_row_after = cur.execute("SELECT staticData FROM workflow_entity WHERE id = ?", (workflow_id,)).fetchone()
    static_after_raw = (static_row_after["staticData"] or "") if static_row_after else ""
    static_has_number_after = any(variant in static_after_raw for variant in variants)
    conn.close()
    return {
        "workflowId": workflow_id,
        "variants": variants,
        "staticHasNumberBefore": static_has_number_before,
        "executionIdsMatched": len(execution_ids),
        "staticHasNumberAfter": static_has_number_after,
        "executionCountAfter": execution_count_after,
        "exclusiveAllowlistApplied": bool(authorized_number),
    }


def cleanup_n8n_via_docker(
    project_dir: Path,
    number: str,
    workflow_id: str,
    dry_run: bool,
    authorized_number: str | None = None,
):
    cmd = [
        "docker", "run", "--rm",
        "-v", "ai_n8n_data:/data",
        "-v", f"{project_dir}:/work",
        "python:3.11-slim",
        "python", "/work/reset-lead-state.py",
        "cleanup-n8n",
        "--number", number,
        "--workflow-id", workflow_id,
    ]
    if dry_run:
        cmd.append("--dry-run")
    if authorized_number:
        cmd += ["--authorized-number", authorized_number]
    result = run(cmd, capture=True)
    stdout = (result.stdout or "").strip().splitlines()
    payload = json.loads(stdout[-1]) if stdout else {}
    return payload


def verify_n8n_via_docker(project_dir: Path, number: str, workflow_id: str, authorized_number: str | None = None):
    cmd = [
        "docker", "run", "--rm",
        "-v", "ai_n8n_data:/data",
        "-v", f"{project_dir}:/work",
        "python:3.11-slim",
        "python", "/work/reset-lead-state.py",
        "verify-n8n",
        "--number", number,
        "--workflow-id", workflow_id,
    ]
    if authorized_number:
        cmd += ["--authorized-number", authorized_number]
    result = run(cmd, capture=True)
    stdout = (result.stdout or "").strip().splitlines()
    payload = json.loads(stdout[-1]) if stdout else {}
    return payload


def collect_router_cache_keys(db_path: Path, variants: list[str]) -> list[str]:
    patterns = build_like_patterns(variants)
    conn = sqlite3.connect(db_path)
    cur = conn.cursor()
    where_parts = ['CAST("number" AS TEXT) LIKE ?' for _ in patterns]
    normalized_messages = [
        row[0]
        for row in cur.execute(
            f'SELECT DISTINCT normalized_message FROM route_logs WHERE {" OR ".join(where_parts)} AND COALESCE(normalized_message, "") <> ""',
            patterns,
        ).fetchall()
    ]
    conn.close()
    return normalized_messages


def cleanup_router_cache(db_path: Path, normalized_messages: list[str], dry_run: bool):
    conn = sqlite3.connect(db_path)
    cur = conn.cursor()
    before = 0
    after = 0
    if normalized_messages:
        placeholders = ",".join("?" * len(normalized_messages))
        before = cur.execute(
            f"SELECT COUNT(*) FROM response_cache WHERE normalized_message IN ({placeholders})",
            normalized_messages,
        ).fetchone()[0]
        if not dry_run and before:
            cur.execute(
                f"DELETE FROM response_cache WHERE normalized_message IN ({placeholders})",
                normalized_messages,
            )
            conn.commit()
        after = cur.execute(
            f"SELECT COUNT(*) FROM response_cache WHERE normalized_message IN ({placeholders})",
            normalized_messages,
        ).fetchone()[0]
    conn.close()
    return {"matchedKeys": len(normalized_messages), "deletedRows": before, "remainingRows": after}


def configure_router_gate(db_path: Path, authorized_number: str, dry_run: bool, exclusive_allowlist: bool = False):
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    before_allowed = [dict(r) for r in cur.execute(
        "SELECT number, reason, added_at FROM always_allowed_numbers ORDER BY added_at DESC"
    ).fetchall()]
    before_blocked = [dict(r) for r in cur.execute(
        "SELECT number, reason, added_at FROM blocked_numbers WHERE number = ?",
        (authorized_number,),
    ).fetchall()]

    if not dry_run:
        cur.execute("DELETE FROM always_allowed_numbers")
        if exclusive_allowlist:
            cur.execute(
                "INSERT OR REPLACE INTO always_allowed_numbers (number, reason, added_at) VALUES (?, ?, ?)",
                (authorized_number, "controlled_real_test_authorized_only", datetime.utcnow().isoformat() + "Z"),
            )
        cur.execute("DELETE FROM blocked_numbers WHERE number = ?", (authorized_number,))
        conn.commit()

    after_allowed = [dict(r) for r in cur.execute(
        "SELECT number, reason, added_at FROM always_allowed_numbers ORDER BY added_at DESC"
    ).fetchall()]
    after_blocked = [dict(r) for r in cur.execute(
        "SELECT number, reason, added_at FROM blocked_numbers WHERE number = ?",
        (authorized_number,),
    ).fetchall()]
    conn.close()
    return {
        "beforeAllowed": before_allowed,
        "beforeBlockedAuthorized": before_blocked,
        "afterAllowed": after_allowed,
        "afterBlockedAuthorized": after_blocked,
    }


def sanitize_ignored_contacts_files(project_dir: Path, variants: list[str], dry_run: bool):
    result = {}

    txt_path = project_dir / "CHATGPT_MACHINE_LEARNING" / "_AUTO_LISTA_DE_CONTATOS_IGNORADOS.txt"
    txt_info = {"path": str(txt_path), "exists": txt_path.exists(), "matchedLines": 0, "remainingLines": 0}
    if txt_path.exists():
        lines = txt_path.read_text(encoding="utf-8", errors="ignore").splitlines()
        kept = [line for line in lines if not contains_any_variant(line, variants)]
        txt_info["matchedLines"] = len(lines) - len(kept)
        txt_info["remainingLines"] = sum(1 for line in kept if contains_any_variant(line, variants))
        if not dry_run and txt_info["matchedLines"]:
            txt_path.write_text("\n".join(kept) + ("\n" if kept else ""), encoding="utf-8")
            txt_info["remainingLines"] = 0
    result["ignoredContactsMlFile"] = txt_info

    xlsx_path = project_dir / "LISTA_DE_CONTATOS_IGNORADOS.xlsx"
    xlsx_info = {"path": str(xlsx_path), "exists": xlsx_path.exists(), "matchedRows": 0, "remainingRows": 0, "openpyxlAvailable": bool(openpyxl)}
    if xlsx_path.exists() and openpyxl:
        wb = openpyxl.load_workbook(xlsx_path)
        total_removed = 0
        total_remaining = 0
        for ws in wb.worksheets:
            rows_to_delete = []
            for row_idx in range(1, ws.max_row + 1):
                values = [str(cell.value or "") for cell in ws[row_idx]]
                if any(contains_any_variant(value, variants) for value in values):
                    rows_to_delete.append(row_idx)
            total_removed += len(rows_to_delete)
            if not dry_run:
                for row_idx in reversed(rows_to_delete):
                    ws.delete_rows(row_idx, 1)
            for row_idx in range(1, ws.max_row + 1):
                values = [str(cell.value or "") for cell in ws[row_idx]]
                if any(contains_any_variant(value, variants) for value in values):
                    total_remaining += 1
        if not dry_run and total_removed:
            wb.save(xlsx_path)
            total_remaining = 0
        xlsx_info["matchedRows"] = total_removed
        xlsx_info["remainingRows"] = total_remaining
    result["ignoredContactsWorkbook"] = xlsx_info

    return result


def evolution_count_sql(jid: str, number: str) -> str:
    return f"""
SELECT 'Contact' AS table_name, COUNT(*) AS total FROM "Contact" WHERE COALESCE(CAST("remoteJid" AS text),'') LIKE '%{number}%'
UNION ALL
SELECT 'Chat', COUNT(*) FROM "Chat" WHERE COALESCE(CAST("remoteJid" AS text),'') LIKE '%{number}%'
UNION ALL
SELECT 'IsOnWhatsapp', COUNT(*) FROM "IsOnWhatsapp" WHERE COALESCE(CAST("remoteJid" AS text),'') LIKE '%{number}%'
UNION ALL
SELECT 'MessageUpdate', COUNT(*) FROM "MessageUpdate" WHERE COALESCE(CAST("remoteJid" AS text),'') LIKE '%{jid}%' OR COALESCE(CAST(participant AS text),'') LIKE '%{number}%'
UNION ALL
SELECT 'Message', COUNT(*) FROM "Message" WHERE CAST(key AS text) LIKE '%{jid}%' OR COALESCE(CAST(participant AS text),'') LIKE '%{number}%';
""".strip()


def evolution_cleanup_sql(jid: str, number: str) -> str:
    return f"""
DELETE FROM "MessageUpdate"
WHERE COALESCE(CAST("remoteJid" AS text),'') LIKE '%{jid}%'
   OR COALESCE(CAST(participant AS text),'') LIKE '%{number}%';

DELETE FROM "Message"
WHERE CAST(key AS text) LIKE '%{jid}%'
   OR COALESCE(CAST(participant AS text),'') LIKE '%{number}%';

DELETE FROM "IsOnWhatsapp"
WHERE COALESCE(CAST("remoteJid" AS text),'') LIKE '%{number}%';

DELETE FROM "Chat"
WHERE COALESCE(CAST("remoteJid" AS text),'') LIKE '%{number}%';

DELETE FROM "Contact"
WHERE COALESCE(CAST("remoteJid" AS text),'') LIKE '%{number}%';
""".strip()


def run_psql_sql(sql: str):
    cmd = ["docker", "exec", "-i", "evolution-postgres", "psql", "-U", "evolution", "-d", "evolution"]
    return subprocess.run(cmd, input=sql, text=True, check=True, capture_output=True)


def parse_count_output(output: str):
    counts = {}
    for line in (output or "").splitlines():
        line = line.strip()
        if not line or line.startswith("(") or line.startswith("table_name") or line.startswith("-"):
            continue
        if "|" not in line:
            continue
        parts = [p.strip() for p in line.split("|")]
        if len(parts) != 2:
            continue
        name, value = parts
        if value.isdigit():
            counts[name] = int(value)
    return counts


def evolution_counts(number: str):
    jid = f"{number}@s.whatsapp.net"
    result = run_psql_sql(evolution_count_sql(jid, number))
    return parse_count_output(result.stdout)


def evolution_cleanup(number: str, dry_run: bool):
    before = evolution_counts(number)
    if not dry_run:
        run_psql_sql(evolution_cleanup_sql(f"{number}@s.whatsapp.net", number))
    after = evolution_counts(number)
    return {"before": before, "after": after}


def restart_services():
    run(["docker", "restart", "n8n", "evolution"])


def verify_local(db_path: Path, variants: list[str], exclude_tables: set[str] | None = None):
    patterns = build_like_patterns(variants)
    conn = sqlite3.connect(db_path)
    cur = conn.cursor()
    tables = [r[0] for r in cur.execute("SELECT name FROM sqlite_master WHERE type='table' ORDER BY name").fetchall()]
    matches = {}
    excluded = exclude_tables or set()
    for table in tables:
        if table in excluded:
            continue
        info = cur.execute(f'PRAGMA table_info("{table}")').fetchall()
        text_cols = [r[1] for r in info if ((r[2] or "").upper() in ("TEXT", "") or "CHAR" in (r[2] or "").upper() or "CLOB" in (r[2] or "").upper())]
        if not text_cols:
            continue
        where = " OR ".join([f'CAST("{col}" AS TEXT) LIKE ?' for col in text_cols for _ in patterns])
        params = [pattern for _ in text_cols for pattern in patterns]
        count = cur.execute(f'SELECT COUNT(*) FROM "{table}" WHERE {where}', params).fetchone()[0]
        if count:
            matches[table] = count
    conn.close()
    return matches


def main_reset(args):
    number = normalize_number(args.number)
    variants = build_number_variants(number)
    project_dir = Path(args.project_dir)
    runtime_root = Path(args.runtime_root)
    backup_dir = runtime_root / "dados" / "backups"
    ensure_dir(backup_dir)
    ts = datetime.now().strftime("%Y%m%d-%H%M%S")

    result = {
        "number": number,
        "variants": variants,
        "dryRun": bool(args.dry_run),
        "backups": {},
    }

    if not args.skip_backup:
        result["backups"]["router"] = backup_file(runtime_root / "dados" / "router_runtime.sqlite", backup_dir, "router_runtime", ts)
        result["backups"]["crm"] = backup_file(project_dir / "crm_operacional.sqlite", backup_dir, "crm_operacional", ts)
        result["backups"]["n8n"] = backup_n8n_db(project_dir, backup_dir, ts)
        try:
            result["backups"]["evolution"] = backup_evolution_db(backup_dir, ts)
        except Exception as exc:
            result["backups"]["evolutionError"] = str(exc)

    result["crm"] = cleanup_sqlite_rows_by_columns(
        project_dir / "crm_operacional.sqlite",
        {
            "leads": ["number", "informed_phone", "whatsapp_id", "remote_jid", "contact_key"],
            "interactions": ["number", "contact_key", "remote_jid", "text"],
            "learning_backlog": ["number", "customer_question", "model_raw"],
            "ignored_contacts_registry": ["number", "raw_value", "contact_name"],
        },
        variants,
        args.dry_run,
    )
    router_db = runtime_root / "dados" / "router_runtime.sqlite"
    cache_keys = collect_router_cache_keys(router_db, variants)
    result["router"] = cleanup_sqlite_rows_by_columns(
        router_db,
        {
            "route_logs": ["number", "push_name", "inbound_text", "normalized_message"],
            "conversation_history": ["contact_key", "message_text"],
            "lead_memory": ["contact_key", "customer_name", "last_inbound_text", "last_outbound_text", "summary", "answered_slots", "open_question"],
            "learning_events": ["contact_key", "inbound_text", "reply_text", "structured_data", "memory_update"],
            "lid_mappings": ["phone_number", "resolved_jid", "remote_jid", "message_id", "push_name"],
            "always_allowed_numbers": ["number"],
            "blocked_numbers": ["number"],
        },
        variants,
        args.dry_run,
    )
    result["routerCache"] = cleanup_router_cache(router_db, cache_keys, args.dry_run)
    result["gate"] = configure_router_gate(router_db, number, args.dry_run, args.exclusive_allowlist)
    result["n8n"] = cleanup_n8n_via_docker(
        project_dir,
        number,
        args.workflow_id,
        args.dry_run,
        authorized_number=number if args.exclusive_allowlist else None,
    )
    result["files"] = sanitize_ignored_contacts_files(project_dir, variants, args.dry_run)
    result["evolution"] = evolution_cleanup(number, args.dry_run)

    if not args.dry_run and not args.skip_restart:
        restart_services()

    result["verify"] = {
        "crm": verify_local(project_dir / "crm_operacional.sqlite", variants),
        "router": verify_local(router_db, variants, exclude_tables={"always_allowed_numbers", "blocked_numbers"}),
        "n8n": verify_n8n_via_docker(
            project_dir,
            number,
            args.workflow_id,
            authorized_number=number if args.exclusive_allowlist else None,
        ),
        "evolution": evolution_counts(number),
    }

    print(json.dumps(result, ensure_ascii=False, indent=2))


def main():
    parser = argparse.ArgumentParser()
    sub = parser.add_subparsers(dest="command")

    sub_cleanup_n8n = sub.add_parser("cleanup-n8n")
    sub_cleanup_n8n.add_argument("--number", required=True)
    sub_cleanup_n8n.add_argument("--workflow-id", default=DEFAULT_WORKFLOW_ID)
    sub_cleanup_n8n.add_argument("--dry-run", action="store_true")
    sub_cleanup_n8n.add_argument("--authorized-number")

    sub_verify_n8n = sub.add_parser("verify-n8n")
    sub_verify_n8n.add_argument("--number", required=True)
    sub_verify_n8n.add_argument("--workflow-id", default=DEFAULT_WORKFLOW_ID)
    sub_verify_n8n.add_argument("--authorized-number")

    parser.add_argument("--number")
    parser.add_argument("--project-dir", default=str(DEFAULT_PROJECT_DIR))
    parser.add_argument("--runtime-root", default=str(DEFAULT_RUNTIME_ROOT))
    parser.add_argument("--workflow-id", default=DEFAULT_WORKFLOW_ID)
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--skip-backup", action="store_true")
    parser.add_argument("--skip-restart", action="store_true")
    parser.add_argument("--exclusive-allowlist", action="store_true")

    args = parser.parse_args()

    if args.command == "cleanup-n8n":
        payload = cleanup_n8n_internal(
            normalize_number(args.number),
            args.workflow_id,
            dry_run=args.dry_run,
            authorized_number=normalize_number(args.authorized_number) if args.authorized_number else None,
        )
        print(json.dumps(payload, ensure_ascii=False))
        return

    if args.command == "verify-n8n":
        payload = cleanup_n8n_internal(
            normalize_number(args.number),
            args.workflow_id,
            dry_run=True,
            authorized_number=normalize_number(args.authorized_number) if args.authorized_number else None,
        )
        print(json.dumps(payload, ensure_ascii=False))
        return

    if not args.number:
        parser.error("--number is required")

    main_reset(args)


if __name__ == "__main__":
    main()
